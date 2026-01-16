"""
Excel 生成服务层
"""
import json
import logging
from pathlib import Path
from typing import Optional, List, Tuple
from uuid import UUID
from html.parser import HTMLParser
from io import StringIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell

from app.models.task import Task, TaskStatus
from app.services.task_service import TaskService
from app.core.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()


class HTMLTableParser(HTMLParser):
    """HTML 表格解析器"""
    
    def __init__(self):
        super().__init__()
        self.in_table = False
        self.in_row = False
        self.in_cell = False
        self.current_cell_content = []
        self.current_row = []
        self.rows = []
        self.cell_tag = None  # 'td' or 'th'
        self.rowspan = 1
        self.colspan = 1
        
    def handle_starttag(self, tag, attrs):
        if tag == 'table':
            self.in_table = True
            self.rows = []
        elif tag == 'tr' and self.in_table:
            self.in_row = True
            self.current_row = []
        elif tag in ('td', 'th') and self.in_row:
            self.in_cell = True
            self.cell_tag = tag
            self.current_cell_content = []
            # 获取 rowspan 和 colspan
            attrs_dict = dict(attrs)
            self.rowspan = int(attrs_dict.get('rowspan', '1'))
            self.colspan = int(attrs_dict.get('colspan', '1'))
    
    def handle_endtag(self, tag):
        if tag == 'table':
            self.in_table = False
        elif tag == 'tr' and self.in_row:
            self.in_row = False
            if self.current_row:
                self.rows.append(self.current_row)
        elif tag in ('td', 'th') and self.in_cell:
            self.in_cell = False
            cell_text = ''.join(self.current_cell_content).strip()
            # 存储单元格内容及其跨度信息
            cell_info = {
                'text': cell_text,
                'rowspan': self.rowspan,
                'colspan': self.colspan,
                'is_header': self.cell_tag == 'th'
            }
            self.current_row.append(cell_info)
            # 重置
            self.rowspan = 1
            self.colspan = 1
    
    def handle_data(self, data):
        if self.in_cell:
            self.current_cell_content.append(data)
    
    def get_table_data(self) -> List[List[dict]]:
        """获取解析后的表格数据"""
        return self.rows


class ExcelService:
    """Excel 生成服务类"""
    
    @staticmethod
    def parse_html_table(html_content: str) -> pd.DataFrame:
        """
        解析 HTML 表格为 DataFrame
        
        Args:
            html_content: HTML 表格内容
            
        Returns:
            DataFrame: 表格数据
        """
        parser = HTMLTableParser()
        parser.feed(html_content)
        rows_data = parser.get_table_data()
        
        if not rows_data:
            return pd.DataFrame()
        
        # 处理合并单元格
        # 首先计算最大列数
        max_cols = 0
        for row in rows_data:
            col_count = sum(cell['colspan'] for cell in row)
            max_cols = max(max_cols, col_count)
        
        # 创建二维数组来存储展开后的数据
        expanded_rows = []
        rowspan_tracker = {}  # 跟踪需要向下合并的单元格 {col_idx: (remaining_rows, text)}
        
        for row in rows_data:
            expanded_row = [''] * max_cols
            col_idx = 0
            
            # 首先处理之前行的 rowspan
            for check_col in range(max_cols):
                if check_col in rowspan_tracker:
                    remaining, text = rowspan_tracker[check_col]
                    expanded_row[check_col] = text
                    remaining -= 1
                    if remaining > 0:
                        rowspan_tracker[check_col] = (remaining, text)
                    else:
                        del rowspan_tracker[check_col]
            
            # 处理当前行的单元格
            for cell in row:
                # 找到下一个空位置
                while col_idx < max_cols and expanded_row[col_idx] != '':
                    col_idx += 1
                
                if col_idx >= max_cols:
                    break
                
                text = cell['text']
                colspan = cell['colspan']
                rowspan = cell['rowspan']
                
                # 填充 colspan
                for i in range(colspan):
                    if col_idx + i < max_cols:
                        expanded_row[col_idx + i] = text if i == 0 else ''
                
                # 如果有 rowspan，记录到 tracker 中
                if rowspan > 1:
                    for i in range(colspan):
                        if col_idx + i < max_cols:
                            rowspan_tracker[col_idx + i] = (rowspan - 1, text if i == 0 else '')
                
                col_idx += colspan
            
            expanded_rows.append(expanded_row)
        
        # 创建 DataFrame
        df = pd.DataFrame(expanded_rows)
        
        return df
    
    @staticmethod
    def extract_tables_from_ocr_json(ocr_json_path: str) -> List[pd.DataFrame]:
        """
        从 OCR JSON 文件中提取所有表格
        
        Args:
            ocr_json_path: OCR JSON 文件路径
            
        Returns:
            List[DataFrame]: 表格列表
        """
        try:
            with open(ocr_json_path, 'r', encoding='utf-8') as f:
                ocr_data = json.load(f)
            
            tables = []
            
            # 遍历所有页面
            pages = ocr_data.get('pages', [])
            for page in pages:
                parsing_res_list = page.get('parsing_res_list', [])
                
                # 提取所有表格
                for block in parsing_res_list:
                    if block.get('block_label') == 'table':
                        block_content = block.get('block_content', '')
                        if block_content:
                            # 清理转义的双引号（\" -> "）
                            block_content = block_content.replace('\\"', '"')
                            
                            # 解析 HTML 表格
                            df = ExcelService.parse_html_table(block_content)
                            if not df.empty:
                                tables.append(df)
                                logger.info(f"提取到表格，形状: {df.shape}")
            
            logger.info(f"从 OCR JSON 中共提取到 {len(tables)} 个表格")
            return tables
            
        except Exception as e:
            logger.error(f"提取表格失败: {str(e)}", exc_info=True)
            raise
    
    @staticmethod
    def create_excel_with_merged_cells(
        html_contents: List[str],
        output_path: str
    ) -> str:
        """
        从 HTML 表格内容创建带合并单元格的 Excel 文件
        
        Args:
            html_contents: HTML 表格内容列表
            output_path: 输出文件路径
            
        Returns:
            str: Excel 文件路径
        """
        if not html_contents:
            raise ValueError("没有表格数据可以生成 Excel")
        
        # 创建工作簿
        wb = Workbook()
        
        # 移除默认的 Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 为每个 HTML 表格创建一个 Sheet
        for table_idx, html_content in enumerate(html_contents):
            sheet_name = f"Table_{table_idx + 1}"
            ws = wb.create_sheet(title=sheet_name)
            
            # 解析 HTML 表格
            parser = HTMLTableParser()
            parser.feed(html_content)
            rows_data = parser.get_table_data()
            
            if not rows_data:
                continue
            
            # 写入数据并记录合并信息
            current_row = 1
            rowspan_tracker = {}  # {col_idx: (end_row, start_row)}
            
            for row_data in rows_data:
                current_col = 1
                
                for cell_info in row_data:
                    # 跳过被 rowspan 占用的单元格
                    while current_col in rowspan_tracker:
                        end_row, start_row = rowspan_tracker[current_col]
                        if current_row <= end_row:
                            current_col += 1
                        else:
                            del rowspan_tracker[current_col]
                            break
                    
                    text = cell_info['text']
                    colspan = cell_info['colspan']
                    rowspan = cell_info['rowspan']
                    is_header = cell_info['is_header']
                    
                    # 写入单元格值
                    cell = ws.cell(row=current_row, column=current_col)
                    cell.value = text if text else ''
                    
                    # 应用合并
                    if colspan > 1 or rowspan > 1:
                        end_col = current_col + colspan - 1
                        end_row = current_row + rowspan - 1
                        
                        # 合并单元格
                        ws.merge_cells(
                            start_row=current_row,
                            start_column=current_col,
                            end_row=end_row,
                            end_column=end_col
                        )
                        
                        # 记录 rowspan 信息
                        if rowspan > 1:
                            for col_offset in range(colspan):
                                col_idx = current_col + col_offset
                                rowspan_tracker[col_idx] = (end_row, current_row)
                    
                    # 设置样式
                    for r in range(current_row, current_row + rowspan):
                        for c in range(current_col, current_col + colspan):
                            cell = ws.cell(row=r, column=c)
                            
                            # 表头样式
                            if is_header or current_row == 1:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                            
                            # 对齐方式
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            
                            # 边框
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                    
                    current_col += colspan
                
                current_row += 1
            
            # 自动调整列宽
            from openpyxl.cell.cell import MergedCell
            from openpyxl.utils import get_column_letter
            
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # 跳过合并单元格
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                if max_length > 0:
                    adjusted_width = min(max_length + 2, 50)  # 最大宽度 50
                    column_letter = get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存 Excel 文件
        wb.save(output_path)
        logger.info(f"Excel 文件已保存（带合并单元格）: {output_path}")
        
        return output_path
    
    @staticmethod
    async def generate_excel_from_ocr(task_id: UUID) -> Tuple[bool, str, Optional[str]]:
        """
        根据任务的 OCR JSON 生成 Excel 文件
        
        Args:
            task_id: 任务 ID
            
        Returns:
            Tuple[bool, str, Optional[str]]: (成功标志, 消息, Excel 路径)
        """
        try:
            # 1. 获取任务
            task = await TaskService.get_task(task_id)
            if not task:
                return False, f"任务不存在: {task_id}", None
            
            # 2. 检查任务状态（允许 ocr_done / excel_generated / editable）
            if task.status not in [TaskStatus.OCR_DONE, TaskStatus.EXCEL_GENERATED, TaskStatus.EDITABLE]:
                return False, f"任务状态错误: {task.status}，期望 ocr_done / excel_generated / editable", None
            
            # 3. 检查 OCR JSON 路径
            if not task.ocr_json_path:
                return False, "OCR JSON 路径为空", None
            
            ocr_json_path = Path(task.ocr_json_path)
            if not ocr_json_path.exists():
                return False, f"OCR JSON 文件不存在: {ocr_json_path}", None
            
            # 4. 提取 HTML 表格内容
            logger.info(f"开始从 OCR JSON 提取表格: {ocr_json_path}")
            
            # 读取 OCR JSON 并提取 HTML 内容
            with open(ocr_json_path, 'r', encoding='utf-8') as f:
                ocr_data = json.load(f)
            
            html_contents = []
            pages = ocr_data.get('pages', [])
            for page in pages:
                parsing_res_list = page.get('parsing_res_list', [])
                for block in parsing_res_list:
                    if block.get('block_label') == 'table':
                        block_content = block.get('block_content', '')
                        if block_content:
                            # 清理转义的双引号（虽然保存时已清理，但保险起见）
                            block_content = block_content.replace('\\"', '"')
                            html_contents.append(block_content)
                            logger.info(f"提取到表格，HTML 长度: {len(block_content)}")
            
            if not html_contents:
                return False, "未从 OCR JSON 中提取到表格", None
            
            # 5. 生成 Excel（带合并单元格）
            excel_dir = Path(settings.data_dir) / 'excel'
            excel_dir.mkdir(parents=True, exist_ok=True)
            
            excel_filename = f"{task_id}.xlsx"
            excel_path = excel_dir / excel_filename
            
            logger.info(f"开始生成 Excel 文件（带合并单元格）: {excel_path}")
            ExcelService.create_excel_with_merged_cells(html_contents, str(excel_path))
            
            # 6. 更新任务状态
            excel_path_abs = excel_path.resolve()
            task.excel_path = str(excel_path_abs)
            task.status = TaskStatus.EXCEL_GENERATED
            await task.save()
            
            logger.info(f"Excel 生成成功，任务状态已更新为 excel_generated")
            
            return True, f"Excel 生成成功，包含 {len(html_contents)} 个 Sheet", str(excel_path_abs)
            
        except Exception as e:
            error_msg = f"生成 Excel 失败: {str(e)}"
            logger.error(error_msg, exc_info=True)
            
            # 更新任务状态为失败
            try:
                task = await TaskService.get_task(task_id)
                if task:
                    task.status = TaskStatus.EXCEL_FAILED
                    task.error_message = error_msg
                    await task.save()
            except Exception as save_error:
                logger.error(f"更新任务状态失败: {str(save_error)}")
            
            return False, error_msg, None
    
    @staticmethod
    async def generate_excel_from_table_data(task_id: UUID, table_data) -> Tuple[bool, str, Optional[str]]:
        """
        从前端编辑的表格数据生成 Excel 文件
        
        Args:
            task_id: 任务 ID
            table_data: TableDataResponse 对象（编辑后的表格数据）
            
        Returns:
            Tuple[bool, str, Optional[str]]: (成功标志, 消息, Excel 路径)
        """
        try:
            # 1. 获取任务
            task = await TaskService.get_task(task_id)
            if not task:
                return False, f"任务不存在: {task_id}", None
            
            # 2. 创建 Excel 工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认的 Sheet
            
            # 3. 为每个 Sheet 创建工作表
            for sheet_data in table_data.sheets:
                ws = wb.create_sheet(title=sheet_data.sheet_name)
                
                # 4. 写入数据并处理合并单元格
                # 记录已合并的区域，避免重复合并
                merged_cells = set()
                
                for row_idx, row in enumerate(sheet_data.data, start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        # 跳过已被合并覆盖的单元格
                        excel_cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(excel_cell, MergedCell):
                            continue

                        # 写入单元格内容
                        excel_cell.value = cell.text
                        
                        # 设置样式
                        if cell.is_header:
                            excel_cell.font = Font(bold=True)
                            excel_cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                        
                        excel_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        excel_cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000')
                        )
                        
                        # 处理合并单元格
                        if (cell.rowspan > 1 or cell.colspan > 1):
                            start_cell = f"{get_column_letter(col_idx)}{row_idx}"
                            end_cell = f"{get_column_letter(col_idx + cell.colspan - 1)}{row_idx + cell.rowspan - 1}"
                            merge_range = f"{start_cell}:{end_cell}"
                            
                            if merge_range not in merged_cells:
                                try:
                                    ws.merge_cells(merge_range)
                                    merged_cells.add(merge_range)
                                except Exception as e:
                                    logger.warning(f"合并单元格失败 {merge_range}: {e}")
                
                # 5. 调整列宽
                for col_idx in range(1, ws.max_column + 1):
                    max_length = 0
                    for row_idx in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell, MergedCell):
                            continue
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # 6. 保存 Excel 文件
            excel_dir = Path(settings.data_dir) / 'excel'
            excel_dir.mkdir(parents=True, exist_ok=True)
            
            excel_filename = f"{task_id}.xlsx"
            excel_path = excel_dir / excel_filename
            
            wb.save(str(excel_path))
            logger.info(f"Excel 生成成功: {excel_path}")
            
            # 7. 更新任务
            excel_path_abs = excel_path.resolve()
            task.excel_path = str(excel_path_abs)
            await task.save()
            
            return True, f"Excel 生成成功，包含 {len(table_data.sheets)} 个 Sheet", str(excel_path_abs)
            
        except Exception as e:
            error_msg = f"生成 Excel 失败: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return False, error_msg, None